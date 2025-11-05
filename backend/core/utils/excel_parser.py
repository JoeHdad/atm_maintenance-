"""
Generic Excel Parser Utility

This module provides functionality to parse any valid Excel file without enforcing
specific column schemas or row formats. It only validates that the file is a valid
Excel file and can be read successfully.
"""

import openpyxl
from datetime import date, datetime, time
from typing import Dict, List, Tuple


class ExcelParserError(Exception):
    """Custom exception for Excel parsing errors."""
    pass


class GenericExcelParser:
    """
    Generic parser for any Excel file.
    
    This parser accepts any valid .xlsx or .xls file without enforcing
    specific column headers or row formats. It only validates that:
    1. The file has a valid Excel extension (.xlsx or .xls)
    2. The file is a valid, readable Excel file (not corrupted)
    3. The file contains at least one worksheet with data
    """
    
    ALLOWED_EXTENSIONS = ['.xlsx', '.xls']
    
    def __init__(self, file_path_or_obj):
        """
        Initialize parser with file path or file object.
        
        Args:
            file_path_or_obj: Path to Excel file or file-like object
        """
        self.file = file_path_or_obj
        self.workbook = None
        self.worksheet = None
        
    def validate_file_extension(self, filename: str) -> bool:
        """
        Validate that the file has an allowed extension.
        
        Args:
            filename: Name of the file
            
        Returns:
            True if valid, False otherwise
        """
        return any(filename.lower().endswith(ext) for ext in self.ALLOWED_EXTENSIONS)
    
    def load_workbook(self):
        """
        Load the Excel workbook.
        
        Raises:
            ExcelParserError: If workbook cannot be loaded or is invalid
        """
        try:
            self.workbook = openpyxl.load_workbook(self.file, data_only=True)
            if not self.workbook.sheetnames:
                raise ExcelParserError("Excel file contains no worksheets")
            self.worksheet = self.workbook.active
        except openpyxl.utils.exceptions.InvalidFileException:
            raise ExcelParserError("Invalid Excel file format. File may be corrupted or not a valid Excel file.")
        except Exception as e:
            raise ExcelParserError(f"Failed to load Excel file: {str(e)}")
    
    def extract_all_data(self) -> List[Dict]:
        """
        Extract all data from the worksheet without schema validation.
        
        Returns:
            List of dictionaries where each dict represents a row
            Keys are column indices (0, 1, 2, ...) or header names if present
            
        Raises:
            ExcelParserError: If worksheet is not loaded
        """
        if not self.worksheet:
            raise ExcelParserError("Worksheet not loaded")
        
        all_data = []
        
        # Get all rows
        for row_num, row in enumerate(self.worksheet.iter_rows(values_only=True), start=1):
            # Skip completely empty rows
            if not any(row):
                continue
            
            # Convert row to dictionary with column indices as keys
            row_dict = {
                f"col_{i}": self._normalize_cell_value(cell_value)
                for i, cell_value in enumerate(row, start=1)
            }
            
            all_data.append(row_dict)
        
        # Check if we have any data
        if not all_data:
            raise ExcelParserError("Excel file contains no data rows")
        
        return all_data

    def _normalize_cell_value(self, value):
        """Convert non-JSON-serializable values to string representations."""
        if isinstance(value, datetime):
            return value.isoformat()
        if isinstance(value, date):
            return value.isoformat()
        if isinstance(value, time):
            return value.isoformat()
        return value
    
    def parse(self) -> List[Dict]:
        """
        Main parsing method that loads and extracts all data from the Excel file.
        
        Returns:
            List of dictionaries containing all data from the worksheet
            
        Raises:
            ExcelParserError: If file cannot be loaded or is invalid
        """
        self.load_workbook()
        all_data = self.extract_all_data()
        
        # Close workbook
        if self.workbook:
            self.workbook.close()
        
        return all_data


def parse_excel_file(file_obj, filename: str) -> Tuple[List[Dict], int]:
    """
    Generic function to parse any Excel file.
    
    This function accepts any valid Excel file without enforcing specific
    column schemas or row formats.
    
    Args:
        file_obj: File-like object containing Excel data
        filename: Name of the file (for extension validation)
        
    Returns:
        Tuple of (parsed_data, row_count) where:
        - parsed_data: List of dictionaries, each representing a row
        - row_count: Number of data rows extracted
        
    Raises:
        ExcelParserError: If file extension is invalid or file cannot be parsed
    """
    parser = GenericExcelParser(file_obj)
    
    # Validate file extension
    if not parser.validate_file_extension(filename):
        raise ExcelParserError(
            f"Invalid file extension. Allowed: {', '.join(parser.ALLOWED_EXTENSIONS)}"
        )
    
    # Parse the file
    parsed_data = parser.parse()
    
    return parsed_data, len(parsed_data)
